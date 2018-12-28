#OpenPyXL is a library used to read and write Excel 2010 xlsx/xlsm/xltx/xltm files
#wget https://www.dropbox.com/s/jhlsih6zs5rz3lr/sample.xlsx?dl=1 to work with sample material


import openpyxl


#Opening an Excel Document
excel_document = openpyxl.load_workbook('sample.xlsx')
print(type(excel_document))
#<class 'openpyxl.workbook.workbook.Workbook'>
# the object returned is Workbook, of data type workbook; the Workbook object here represents the Excel file


#Sheets (sheets consist of columns A,B,C,etc. and rows 1,2,3,4,etc.)
print(excel_document.get_sheet_names()) # we have one sheet with name [u'Sheet1']
#If you have multiple sheets, you can access a specific sheet by its name using get_sheet_by_name()


#Accessing Cells
sheet = excel_document.get_sheet_by_name('Sheet1')
print(sheet['A2'].value) # value of cell A2 is Abder
print(sheet.cell(row = 5, column = 2).value) #for priting using rown-column notation

#for printing type of cell we can do:
print(type(sheet['A2']))
print(sheet.cell(row = 5, column = 2))


#Accessing a Range of Cells (say from A1 to B3)
multiple_cells = sheet['A1':'B3']
for row in multiple_cells:
    for cell in row:
        print(cell.value)
#Should retunr:
#Name
#Profession
#Abder
#Student
#Bob
#Engineer


#Extracting all rows and columns - TODO: fix this using https://stackoverflow.com/questions/44257659/typeerror-generator-object-is-not-subscriptable-csv-file

#for rows:
all_rows = sheet.rows
print(all_rows[:])
#returns a tupple: ((<Cell Sheet1.A1>, <Cell Sheet1.B1>), (<Cell Sheet1.A2>, <Cell Sheet1.B2>), (<Cell Sheet1.A3>, <Cell Sheet1.B3>), (<Cell Sheet1.A4>, <Cell Sheet1.B4>), (<Cell Sheet1.A5>, <Cell Sheet1.B5>), (<Cell Sheet1.A6>, <Cell Sheet1.B6>), (<Cell Sheet1.A7>, <Cell Sheet1.B7>))

#for columns"
all_columns = sheet.columns
print(all_columns[:])
#returns a tupple: ((<Cell Sheet1.A1>, <Cell Sheet1.A2>, <Cell Sheet1.A3>, <Cell Sheet1.A4>, <Cell Sheet1.A5>, <Cell Sheet1.A6>, <Cell Sheet1.A7>), (<Cell Sheet1.B1>, <Cell Sheet1.B2>, <Cell Sheet1.B3>, <Cell Sheet1.B4>, <Cell Sheet1.B5>, <Cell Sheet1.B6>, <Cell Sheet1.B7>))


#find out more at official dicumentation https://openpyxl.readthedocs.org/en/default/index.html
